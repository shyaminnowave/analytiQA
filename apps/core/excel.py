from celery import shared_task
from django.contrib.contenttypes.models import ContentType
from openpyxl import load_workbook
from apps.account.models import Account
from apps.core.models import (
    TestCaseModel,
    Tag
)
from apps.core.utlity import QuerySetEntry
from abc import ABC, abstractmethod
from collections import defaultdict
from rest_framework.status import HTTP_200_OK, HTTP_400_BAD_REQUEST
from apps.general.models import Notification
from collections import defaultdict


class FileFactory(ABC):

    @abstractmethod
    def import_data(self):
        pass

class ExcelFileFactory(FileFactory):
    """Base factory class to handle Excel file operations."""

    def __init__(self, file, user):
        self.response_format = {
            "status": True,
            "status_code": HTTP_200_OK,
            "data": "",
            "message": ""
        }
        self.file = file
        self.user = user

    def _init_workbook(self):
        """Initialize the workbook and return the active sheet."""
        workbook = load_workbook(self.file)
        return workbook.active

    def import_data(self):
        pass


class TestCaseExl(ExcelFileFactory):

    def __init__(self, file, user):
        super(TestCaseExl, self).__init__(file, user)
        self.ws = self._init_workbook()

    def get_tag(self, tags):
        _ins = []
        for tg in tags:
            tag_instance, created = Tag.objects.get_or_create(name=tg)
            _ins.append(tag_instance)
        return _ins
    
    def _build_error_response(self, error):
        self.response_format.update({
            "status": False,
            "status_code": HTTP_400_BAD_REQUEST,
            "message": str(error),
        })
        return self.response_format

    def _build_success_response(self, message):
        self.response_format.update({
            "data": "Success",
            "message": message,
        })
        return self.response_format

    def get_row_dict(self):
        sheet = self.ws
        headings = {}
        for idx, col in enumerate(sheet.iter_cols(min_row=1, max_row=1, values_only=False)):
            for cell in col:
                headings[cell.value] = idx
        return headings

    def _parse_step(self):
        """
        Parse the step data from the worksheet using your improved logic
        """
        sheet = self.ws
        data = []
        current_key = None
        prev = None
        headings = self.get_row_dict()

        print(headings)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = row[headings['Key']] if row[headings['Key']] is not None else current_key
            value = row[headings['Manual Test Steps'] + 1]
            print(key, value)
            if key:
                current_key = key
            prev = value if value in ['Action', 'Data', 'Expected Result'] else prev
            if value != None and value not in ["Action", "Data", "Expected Result"]:
                _ts = {
                    prev: value
                }
                data.append((current_key, _ts))

        test_cases = defaultdict(dict)
        current = {}
        step_counter = defaultdict(int)
        prev_case_id = None

        for case_id, detail in data:
            key = list(detail.keys())[0]
            value = list(detail.values())[0]

            if case_id != prev_case_id and current:
                step_counter[prev_case_id] += 1
                test_cases[prev_case_id][step_counter[prev_case_id]] = current
                current = {}

            if key == 'Action':
                if current:
                    step_counter[case_id] += 1
                    test_cases[case_id][step_counter[case_id]] = current
                current = {'step_action': value, 'step_data': '', 'expected_result': ''}
            elif key == 'Data':
                current['step_data'] = value
            elif key == 'Expected Result':
                current['expected_result'] = f"{current['expected_result']} {value}" if current[
                    'expected_result'] else value

            prev_case_id = case_id

        if current:
            step_counter[prev_case_id] += 1
            test_cases[prev_case_id][step_counter[prev_case_id]] = current
        print(dict(test_cases))
        return dict(test_cases)

    def import_data(self):
        """
        Import TestCase and Its Related Data from the Excel
        """
        testcase_list = set(TestCaseModel.objects.values_list('jira_id', flat=True))
        _steps = self._parse_step()
        headings = self.get_row_dict()
        testcases_with_tags = []
        tests = []
        _step = dict()
        status = {
            "To Do": "todo",
            "Completed": "completed",
            "Ongoing": "ongoing"
        }
        types = {
            "Performance": "performance",
            "Soak": "soak",
            "Smoke": "smoke"
        }
        priority = {
            "Class 1": "class_1",
            "Class 2": "class_2",
            "Class 3": "class_3"
        }
        try:
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row[headings['Key']] is not None:
                    jira_id = str(row[headings['Key']]).split("-")[1] if row[headings['Key']] else None
                    if int(jira_id) not in testcase_list:
                        tag_names = [tag.strip() for tag in row[headings['Labels']].split(",")] if row[headings['Labels']] else []
                        _data = {
                            "jira_id": jira_id,
                            "name": row[headings['Summary']] if row[headings['Summary']] is not None else None,
                            "summary": row[headings['Summary']],
                            "description": row[headings['Summary']],
                            "priority": priority.get(row[headings['Priority']], 'class_3'),
                            "testcase_type": 'performance',
                            "status": status.get(row[headings['Status']], 'todo'),
                            "reporter": row[headings['Reporter']],
                            "created_by": self.user,
                            "steps": dict(_steps[row[headings['Key']]]),
                        }
                        tests.append(_data)
                        testcases_with_tags.append({int(str(row[headings['Key']]).split("-")[1]): tag_names})
            if tests:
                try:
                    QuerySetEntry.bulk_create_entries(TestCaseModel, [TestCaseModel(**i) for i in tests])
                except Exception as e:
                    print('error', str(e))
                for i in testcases_with_tags:
                    for instance, tag in i.items():
                        _instance = TestCaseModel.objects.get(jira_id=instance)
                        tags = self.get_tag(tag)
                        _instance.tags.set(tags)
                Notification.objects.create(
                    message=f"Testcases Uploaded Successfully!",
                    user=self.user if self.user else None,
                    assigned_to=self.user if self.user else None,
                    content_type=ContentType.objects.get_for_model(TestCaseModel),
                    status=True
                )
            else:
                return self._build_success_response("TestCases Already Present in the DataBase")
        except Exception as e:
            Notification.objects.create(
                message=f"Testcases Uploaded Failed!",
                user=self.user if self.user else None,
                assigned_to=self.user if self.user else None,
                content_type=ContentType.objects.get_for_model(TestCaseModel),
                status=False
            )
            print(str(e))
            return self._build_error_response(e)
        return self._build_success_response("TestCase Uploaded Successfully")

